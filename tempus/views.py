from django.contrib import messages
from django.contrib.auth import authenticate, get_user_model, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from datetime import date as dt_date
from django.utils import timezone
from django.utils.formats import date_format
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_http_methods, require_POST
from io import BytesIO
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from tempus.models import Ambiente, ContractorProfile, Ficha, MatrixResult, MatrixUpload, PlantProfile, ProfileAvatar, Sede
from tempus.services.plant_users_import import (
    parse_plant_users_workbook,
    save_plant_users_summary,
)
from tempus.services.matrix_import import parse_matrix_workbook, save_matrix_summary
from tempus.services.users_import import parse_users_workbook, save_users_summary


def _is_admin(user) -> bool:
    return user.is_staff or user.is_superuser


def _resolve_role(user) -> str:
    if user.is_superuser or user.is_staff:
        return "Administrador"

    if user.groups.filter(name="Planta").exists():
        return "Planta"

    if user.groups.filter(name="Contratista").exists():
        return "Contratista"

    return "Usuario"


@never_cache
@require_http_methods(["GET", "POST"])
def login_view(request):
    if request.user.is_authenticated:
        return redirect("dashboard")

    if request.method == "POST":
        correo = request.POST.get("correo", "").strip().lower()
        password = request.POST.get("password", "")

        username = correo
        if "@" in correo:
            user_model = get_user_model()
            user = user_model.objects.filter(email__iexact=correo).first()
            if user:
                username = user.get_username()

        user = authenticate(request, username=username, password=password)

        if user is None:
            messages.error(request, "Credenciales invalidas.")
            return render(request, "login.html", status=401)

        if not (user.is_staff or user.is_superuser):
            messages.error(request, "Tu usuario no tiene rol admin.")
            return render(request, "login.html", status=403)

        login(request, user)
        return redirect("dashboard")

    return render(request, "login.html")


@never_cache
@login_required
def dashboard_view(request):
    if not _is_admin(request.user):
        logout(request)
        messages.error(request, "Tu sesion no tiene permisos admin.")
        return redirect("login")

    return render(request, "dashboard.html", {"active_tab": "inicio"})


@never_cache
@login_required
@require_http_methods(["GET", "POST"])
def profile_view(request):
    user = request.user
    contractor = getattr(user, "contractor_profile", None)
    plant = getattr(user, "plant_profile", None)
    avatar = getattr(user, "profile_avatar", None)
    last_login_display = "-"

    if user.last_login:
        last_login_display = date_format(
            timezone.localtime(user.last_login),
            "j \\d\\e F \\d\\e Y, g:i A",
        )

    if request.method == "POST":
        action = request.POST.get("action", "info")

        if action == "info":
            first_name = request.POST.get("first_name", "").strip()
            last_name = request.POST.get("last_name", "").strip()
            email = request.POST.get("email", "").strip()
            user.first_name = first_name
            user.last_name = last_name
            if email:
                user.email = email
            user.save(update_fields=["first_name", "last_name", "email"])
            messages.success(request, "Información actualizada correctamente.")

        elif action == "avatar":
            avatar_file = request.FILES.get("avatar")
            if not avatar_file:
                messages.error(request, "Debes seleccionar una imagen para actualizar la foto.")
            elif avatar_file.size > 5 * 1024 * 1024:
                messages.error(request, "La imagen supera 5MB. Usa un archivo mas ligero.")
            elif not str(avatar_file.content_type or "").startswith("image/"):
                messages.error(request, "El archivo debe ser una imagen valida.")
            else:
                avatar_obj, _ = ProfileAvatar.objects.get_or_create(user=user)
                avatar_obj.image = avatar_file
                avatar_obj.save(update_fields=["image", "updated_at"])
                avatar = avatar_obj
                messages.success(request, "Foto de perfil actualizada correctamente.")

        elif action == "password":
            current = request.POST.get("current_password", "")
            new_pass = request.POST.get("new_password", "")
            confirm = request.POST.get("confirm_password", "")
            if not user.check_password(current):
                messages.error(request, "La contraseña actual es incorrecta.")
            elif len(new_pass) < 8:
                messages.error(request, "La nueva contraseña debe tener al menos 8 caracteres.")
            elif new_pass != confirm:
                messages.error(request, "Las contraseñas nuevas no coinciden.")
            else:
                user.set_password(new_pass)
                user.save()
                update_session_auth_hash(request, user)
                messages.success(request, "Contraseña actualizada correctamente.")

    context = {
        "active_tab": "perfil",
        "contractor": contractor,
        "plant": plant,
        "last_login_display": last_login_display,
        "avatar_url": avatar.image.url if avatar and avatar.image else "",
    }
    return render(request, "profile.html", context)


@never_cache
@login_required
@require_http_methods(["GET", "POST"])
def matrix_upload_view(request):
    if request.method == "GET":
        request.session.pop("matrix_preview", None)

    preview = request.session.get("matrix_preview")

    if request.method == "POST":
        action = request.POST.get("action", "preview")

        if action == "save":
            if not preview:
                messages.error(request, "Primero debes procesar una matriz antes de guardarla.")
            else:
                try:
                    save_matrix_summary(preview, request.user)
                    request.session.pop("matrix_preview", None)
                    preview = None
                    messages.success(request, "Matriz guardada correctamente en la BD local.")
                except ValueError as error:
                    messages.error(request, str(error))
        else:
            uploaded_file = request.FILES.get("matrix_file")
            if not uploaded_file:
                messages.error(request, "Debes seleccionar un archivo Excel.")
            else:
                try:
                    summary = parse_matrix_workbook(uploaded_file.read())
                    preview = {
                        "project_code": summary.project_code,
                        "project_description": summary.project_description,
                        "program_code": summary.program_code,
                        "program_name": summary.program_name,
                        "level": summary.level,
                        "version": summary.version,
                        "competencies": summary.competencies,
                        "rows_processed": summary.rows_processed,
                        "skipped_rows": summary.skipped_rows,
                        "competency_count": len(summary.competencies),
                        "result_count": sum(len(competency.get("results", [])) for competency in summary.competencies),
                    }
                    preview["project_code_exists"] = MatrixUpload.objects.filter(project_code=summary.project_code).exists()
                    request.session["matrix_preview"] = preview
                    messages.success(request, "Archivo procesado. Revisa la previsualizacion antes de guardarlo.")
                except Exception:
                    messages.error(request, "No se pudo leer el archivo. Verifica que sea un Excel valido (.xlsx).")

    return render(
        request,
        "matrix_upload.html",
        {
            "active_tab": "crear",
            "preview": preview,
        },
    )


@never_cache
@login_required
def matrix_detail_view(request, pk):
    if not _is_admin(request.user):
        logout(request)
        messages.error(request, "Tu sesion no tiene permisos admin.")
        return redirect("login")

    matrix = get_object_or_404(
        MatrixUpload.objects.prefetch_related("competencies__results"),
        pk=pk,
    )

    max_trimester = 0
    for competency in matrix.competencies.all():
        for result in competency.results.all():
            try:
                value = int(result.trimester)
                if value > max_trimester:
                    max_trimester = value
            except (ValueError, TypeError):
                pass

    return render(
        request,
        "matrix_detail.html",
        {
            "active_tab": "matrices",
            "matrix": matrix,
            "max_trimester": max_trimester if max_trimester > 0 else None,
        },
    )


@never_cache
@login_required
def matrix_list_view(request):
    if not _is_admin(request.user):
        logout(request)
        messages.error(request, "Tu sesion no tiene permisos admin.")
        return redirect("login")

    all_matrices = MatrixUpload.objects.prefetch_related(
        "competencies__results"
    ).order_by("-created_at")

    matrices_data = []
    for matrix in all_matrices:
        max_trimester = 0
        for competency in matrix.competencies.all():
            for result in competency.results.all():
                try:
                    value = int(result.trimester)
                    if value > max_trimester:
                        max_trimester = value
                except (ValueError, TypeError):
                    pass
        matrices_data.append(
            {
                "matrix": matrix,
                "max_trimester": max_trimester if max_trimester > 0 else None,
            }
        )

    return render(
        request,
        "matrix_list.html",
        {
            "active_tab": "matrices",
            "matrices": matrices_data,
        },
    )


@never_cache
@require_POST
@login_required
def matrix_delete_view(request, pk):
    if not _is_admin(request.user):
        messages.error(request, "No tienes permisos para eliminar matrices.")
        return redirect("matrix-list")

    matrix = get_object_or_404(MatrixUpload, pk=pk)
    program_name = matrix.program_name
    matrix.delete()
    messages.success(request, f"Matriz '{program_name}' eliminada correctamente junto con sus competencias y resultados.")
    return redirect("matrix-list")


@never_cache
@require_POST
@login_required
def logout_view(request):
    logout(request)
    response = redirect("login")
    response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    return response


@never_cache
@login_required
def users_panel_view(request):
    if not _is_admin(request.user):
        logout(request)
        messages.error(request, "Tu sesion no tiene permisos admin.")
        return redirect("login")

    user_model = get_user_model()
    users = list(
        user_model.objects.select_related("contractor_profile", "plant_profile").prefetch_related("groups")
    )

    def sort_key(user):
        role = _resolve_role(user)
        display_name = getattr(getattr(user, "contractor_profile", None), "full_name", "") or user.first_name or user.get_username()

        if user.pk == request.user.pk and role == "Administrador":
            priority = 0
        elif role == "Administrador":
            priority = 1
        else:
            priority = 2

        return (priority, display_name.lower(), user.get_username().lower())

    users.sort(key=sort_key)

    users_data = []
    contractors_count = 0
    plant_count = 0
    admins_count = 0

    for user in users:
        role = _resolve_role(user)
        if role == "Contratista":
            contractors_count += 1
        if role == "Planta":
            plant_count += 1
        if role == "Administrador":
            admins_count += 1

        contractor = getattr(user, "contractor_profile", None)
        plant = getattr(user, "plant_profile", None)
        users_data.append(
            {
                "user": user,
                "role": role,
                "contractor": contractor,
                "plant": plant,
                "is_current_admin": user.pk == request.user.pk and role == "Administrador",
            }
        )

    return render(
        request,
        "users_panel.html",
        {
            "active_tab": "usuarios",
            "users_data": users_data,
            "total_users": len(users),
            "total_contractors": contractors_count,
            "total_plant": plant_count,
            "total_admins": admins_count,
        },
    )


@never_cache
@require_POST
@login_required
def users_delete_non_admin_view(request):
    if not _is_admin(request.user):
        messages.error(request, "No tienes permisos para eliminar usuarios.")
        return redirect("users-panel")

    user_model = get_user_model()
    users_to_delete = user_model.objects.filter(is_staff=False, is_superuser=False)
    deleted_count = users_to_delete.count()

    if deleted_count == 0:
        messages.info(request, "No hay usuarios no administradores para eliminar.")
        return redirect("users-panel")

    users_to_delete.delete()
    messages.success(
        request,
        f"Se eliminaron {deleted_count} usuarios no administradores junto con sus datos relacionados.",
    )
    return redirect("users-panel")


@never_cache
@login_required
@require_http_methods(["GET", "POST"])
def users_upload_view(request):
    if not _is_admin(request.user):
        logout(request)
        messages.error(request, "Tu sesion no tiene permisos admin.")
        return redirect("login")

    if request.method == "GET":
        request.session.pop("users_preview", None)

    preview = request.session.get("users_preview")

    if request.method == "POST":
        action = request.POST.get("action", "preview")

        if action == "save":
            if not preview:
                messages.error(request, "Primero debes cargar y previsualizar un archivo.")
            else:
                try:
                    created_users, updated_users = save_users_summary(preview)
                    request.session.pop("users_preview", None)
                    preview = None
                    messages.success(
                        request,
                        f"Usuarios procesados correctamente. Nuevos: {created_users}. Actualizados: {updated_users}.",
                    )
                except ValueError as error:
                    messages.error(request, str(error))
        else:
            uploaded_file = request.FILES.get("users_file")
            if not uploaded_file:
                messages.error(request, "Debes seleccionar un archivo Excel de contratistas.")
            else:
                try:
                    summary = parse_users_workbook(uploaded_file.read())

                    existing_documents = set(
                        ContractorProfile.objects.filter(
                            document_number__in=[c.get("document_number", "") for c in summary.contractors]
                        ).values_list("document_number", flat=True)
                    )
                    existing_contracts = set(
                        ContractorProfile.objects.filter(
                            contract_number__in=[c.get("contract_number", "") for c in summary.contractors]
                        ).values_list("contract_number", flat=True)
                    )

                    contractors_preview = []
                    for contractor in summary.contractors:
                        will_update = (
                            contractor.get("document_number") in existing_documents
                            or contractor.get("contract_number") in existing_contracts
                        )

                        start_date = contractor.get("contract_start_date")
                        end_date = contractor.get("contract_end_date")

                        contractors_preview.append(
                            {
                                **contractor,
                                "contract_start_date": start_date.isoformat() if hasattr(start_date, "isoformat") and start_date else "",
                                "contract_end_date": end_date.isoformat() if hasattr(end_date, "isoformat") and end_date else "",
                                "will_update": will_update,
                            }
                        )

                    preview = {
                        "rows_processed": summary.rows_processed,
                        "skipped_rows": summary.skipped_rows,
                        "contractors": contractors_preview,
                        "create_count": sum(1 for c in contractors_preview if not c["will_update"]),
                        "update_count": sum(1 for c in contractors_preview if c["will_update"]),
                    }
                    request.session["users_preview"] = preview
                    messages.success(request, "Archivo procesado. Revisa la previsualizacion antes de guardar.")
                except ValueError as error:
                    messages.error(request, str(error))
                except Exception:
                    messages.error(request, "No se pudo leer el archivo. Verifica que sea un Excel valido (.xlsx).")

    return render(
        request,
        "users_upload.html",
        {
            "active_tab": "usuarios",
            "preview": preview,
        },
    )


@never_cache
@login_required
@require_http_methods(["GET"])
def users_upload_template_view(request):
    if not _is_admin(request.user):
        logout(request)
        messages.error(request, "Tu sesion no tiene permisos admin.")
        return redirect("login")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Contratistas"

    sheet.append(
        [
            "NO.",
            "N° CONTRATO",
            "NOMBRE Y APELLIDOS COMPLETOS",
            "No. de Documento de identidad",
            "Correo @SENA [maria@sena.edu.co]",
            "Nivel de Formación Máximo Alcanzado",
            "Pregrado",
            "Posgrado y de mas",
            "Coordinación a la que pertenece",
            "Modalidad de Formación [Gestión]",
            "Especialidad y/o área en la que se desempeña en el CEDE",
            "Fecha inicio contrato (Solo para instructores)",
            "Fecha finalización contrato (Solo para instructores)",
        ]
    )

    # Formato visual de encabezados para una plantilla más clara y legible.
    column_widths = [8, 16, 38, 14, 22, 16, 28, 28, 10, 20, 20, 20, 20]
    header_fill = PatternFill(fill_type="solid", fgColor="67BCE8")
    header_font = Font(name="Calibri", bold=True, color="000000")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_border = Border(
        left=Side(style="thin", color="B7B7B7"),
        right=Side(style="thin", color="B7B7B7"),
        top=Side(style="thin", color="B7B7B7"),
        bottom=Side(style="thin", color="B7B7B7"),
    )

    for col_idx, width in enumerate(column_widths, start=1):
        column_letter = chr(64 + col_idx)
        sheet.column_dimensions[column_letter].width = width
        header_cell = sheet.cell(row=1, column=col_idx)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = header_alignment
        header_cell.border = header_border

    sheet.row_dimensions[1].height = 70
    sheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="plantilla_contratistas.xlsx"'
    return response


@never_cache
@login_required
@require_http_methods(["GET"])
def users_plant_upload_template_view(request):
    if not _is_admin(request.user):
        logout(request)
        messages.error(request, "Tu sesion no tiene permisos admin.")
        return redirect("login")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Planta"

    sheet.append(
        [
            "N°",
            "Nombre y apellidos",
            "CEDULA",
            "AREA",
            "ESTUDIOS",
        ]
    )

    column_widths = [8, 52, 16, 42, 64]
    header_fill = PatternFill(fill_type="solid", fgColor="A9D08E")
    header_font = Font(name="Calibri", bold=True, color="000000")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_border = Border(
        left=Side(style="thin", color="B7B7B7"),
        right=Side(style="thin", color="B7B7B7"),
        top=Side(style="thin", color="B7B7B7"),
        bottom=Side(style="thin", color="B7B7B7"),
    )

    for col_idx, width in enumerate(column_widths, start=1):
        column_letter = chr(64 + col_idx)
        sheet.column_dimensions[column_letter].width = width
        header_cell = sheet.cell(row=1, column=col_idx)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = header_alignment
        header_cell.border = header_border

    sheet.row_dimensions[1].height = 32
    sheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="plantilla_planta.xlsx"'
    return response


@never_cache
@login_required
@require_http_methods(["GET", "POST"])
def users_plant_upload_view(request):
    if not _is_admin(request.user):
        logout(request)
        messages.error(request, "Tu sesion no tiene permisos admin.")
        return redirect("login")

    if request.method == "GET":
        request.session.pop("users_plant_preview", None)

    preview = request.session.get("users_plant_preview")

    if request.method == "POST":
        action = request.POST.get("action", "preview")

        if action == "save":
            if not preview:
                messages.error(request, "Primero debes cargar y previsualizar un archivo.")
            else:
                try:
                    created_users, updated_users = save_plant_users_summary(preview)
                    request.session.pop("users_plant_preview", None)
                    preview = None
                    messages.success(
                        request,
                        f"Instructores planta procesados correctamente. Nuevos: {created_users}. Actualizados: {updated_users}.",
                    )
                except ValueError as error:
                    messages.error(request, str(error))
        else:
            uploaded_file = request.FILES.get("users_file")
            if not uploaded_file:
                messages.error(request, "Debes seleccionar un archivo Excel de instructores planta.")
            else:
                try:
                    summary = parse_plant_users_workbook(uploaded_file.read())

                    existing_documents = set(
                        PlantProfile.objects.filter(
                            document_number__in=[i.get("document_number", "") for i in summary.instructors]
                        ).values_list("document_number", flat=True)
                    )

                    instructors_preview = []
                    for instructor in summary.instructors:
                        will_update = instructor.get("document_number") in existing_documents
                        instructors_preview.append(
                            {
                                **instructor,
                                "will_update": will_update,
                            }
                        )

                    preview = {
                        "rows_processed": summary.rows_processed,
                        "skipped_rows": summary.skipped_rows,
                        "instructors": instructors_preview,
                        "create_count": sum(1 for i in instructors_preview if not i["will_update"]),
                        "update_count": sum(1 for i in instructors_preview if i["will_update"]),
                    }
                    request.session["users_plant_preview"] = preview
                    messages.success(request, "Archivo de planta procesado. Revisa la previsualizacion antes de guardar.")
                except ValueError as error:
                    messages.error(request, str(error))
                except Exception:
                    messages.error(request, "No se pudo leer el archivo. Verifica que sea un Excel valido (.xlsx).")

    return render(
        request,
        "users_upload_plant.html",
        {
            "active_tab": "usuarios",
            "preview": preview,
        },
    )


@never_cache
@login_required
@require_http_methods(["GET", "POST"])
def sites_ambiences_view(request):
    if not _is_admin(request.user):
        logout(request)
        messages.error(request, "Tu sesion no tiene permisos admin.")
        return redirect("login")

    if request.method == "POST":
        action = request.POST.get("action", "")

        if action == "create_sede":
            nombre = request.POST.get("nombre", "").strip()
            ubicacion = request.POST.get("ubicacion", "").strip()

            if not nombre or not ubicacion:
                messages.error(request, "Debes completar nombre y direccion de la sede.")
            elif Sede.objects.filter(nombre__iexact=nombre).exists():
                messages.error(request, "Ya existe una sede con ese nombre.")
            else:
                Sede.objects.create(nombre=nombre, ubicacion=ubicacion)
                messages.success(request, "Sede registrada correctamente.")

        elif action == "update_sede":
            sede_id = request.POST.get("sede_id", "").strip()
            nombre = request.POST.get("nombre", "").strip()
            ubicacion = request.POST.get("ubicacion", "").strip()

            sede = Sede.objects.filter(pk=sede_id).first() if sede_id.isdigit() else None

            if not sede:
                messages.error(request, "No se encontro la sede a editar.")
            elif not nombre or not ubicacion:
                messages.error(request, "Debes completar nombre y direccion de la sede.")
            elif Sede.objects.filter(nombre__iexact=nombre).exclude(pk=sede.pk).exists():
                messages.error(request, "Ya existe otra sede con ese nombre.")
            else:
                sede.nombre = nombre
                sede.ubicacion = ubicacion
                sede.save(update_fields=["nombre", "ubicacion"])
                messages.success(request, "Sede actualizada correctamente.")

        elif action == "create_ambiente":
            sede_id = request.POST.get("sede_id", "").strip()
            nombre = request.POST.get("nombre_ambiente", "").strip()
            descripcion = request.POST.get("descripcion", "").strip()

            sede = Sede.objects.filter(pk=sede_id).first() if sede_id.isdigit() else None

            if not sede:
                messages.error(request, "Debes seleccionar una sede valida.")
            elif not nombre:
                messages.error(request, "Debes ingresar el nombre del ambiente.")
            elif Ambiente.objects.filter(sede=sede, nombre__iexact=nombre).exists():
                messages.error(request, "Ese ambiente ya existe para la sede seleccionada.")
            else:
                Ambiente.objects.create(sede=sede, nombre=nombre, descripcion=descripcion)
                messages.success(request, "Ambiente creado correctamente.")

        elif action == "update_ambiente":
            ambiente_id = request.POST.get("ambiente_id", "").strip()
            sede_id = request.POST.get("sede_id", "").strip()
            nombre = request.POST.get("nombre_ambiente", "").strip()
            descripcion = request.POST.get("descripcion", "").strip()

            ambiente = Ambiente.objects.select_related("sede").filter(pk=ambiente_id).first() if ambiente_id.isdigit() else None
            sede = Sede.objects.filter(pk=sede_id).first() if sede_id.isdigit() else None

            if not ambiente:
                messages.error(request, "No se encontro el ambiente a editar.")
            elif not sede:
                messages.error(request, "Debes seleccionar una sede valida.")
            elif not nombre:
                messages.error(request, "Debes ingresar el nombre del ambiente.")
            elif Ambiente.objects.filter(sede=sede, nombre__iexact=nombre).exclude(pk=ambiente.pk).exists():
                messages.error(request, "Ese ambiente ya existe para la sede seleccionada.")
            else:
                ambiente.sede = sede
                ambiente.nombre = nombre
                ambiente.descripcion = descripcion
                ambiente.save(update_fields=["sede", "nombre", "descripcion"])
                messages.success(request, "Ambiente actualizado correctamente.")

    sedes = Sede.objects.prefetch_related("ambientes").all()
    total_ambientes = sum(sede.ambientes.count() for sede in sedes)

    return render(
        request,
        "sites_ambiences.html",
        {
            "active_tab": "crear",
            "sedes": sedes,
            "total_sedes": sedes.count(),
            "total_ambientes": total_ambientes,
        },
    )


@never_cache
@login_required
@require_http_methods(["GET", "POST"])
def fichas_view(request):
    if not _is_admin(request.user):
        logout(request)
        messages.error(request, "Tu sesion no tiene permisos admin.")
        return redirect("login")

    fichas = Ficha.objects.select_related("matrix").all()

    return render(
        request,
        "fichas_panel.html",
        {
            "active_tab": "crear",
            "fichas": fichas,
            "total_fichas": fichas.count(),
        },
    )


def _get_avatar_url(user):
    """Return profile avatar URL for a user, or empty string if none."""
    try:
        av = user.profile_avatar
        if av.image:
            return av.image.url
    except Exception:
        pass
    return ""


@never_cache
@login_required
@require_http_methods(["GET"])
def ficha_config_view(request, pk):
    if not _is_admin(request.user):
        logout(request)
        messages.error(request, "Tu sesion no tiene permisos admin.")
        return redirect("login")

    ficha = get_object_or_404(Ficha.objects.select_related("matrix", "created_by"), pk=pk)

    raw_results = list(
        MatrixResult.objects
        .filter(competency__matrix=ficha.matrix)
        .select_related("competency")
        .order_by("trimester", "competency__display_order", "display_order")
    )

    cdf_limit = ficha.cdf_trimestres if ficha.cdf_trimestres else None
    results = []
    for result in raw_results:
        if not cdf_limit:
            results.append(result)
            continue

        try:
            trimester_value = int(result.trimester)
        except (TypeError, ValueError):
            # Keep non-numeric rows visible instead of silently dropping them.
            results.append(result)
            continue

        if trimester_value <= cdf_limit:
            results.append(result)

    trimestres = sorted(
        {
            int(r.trimester)
            for r in results
            if r.trimester is not None and str(r.trimester).isdigit()
        }
    )

    # Build unified instructor list (Planta + Contratista)
    instructors = []
    for p in PlantProfile.objects.select_related("user").order_by("full_name"):
        instructors.append({
            "role": "Planta",
            "full_name": p.full_name,
            "document": p.document_number,
            "specialty": (p.studies or p.area or "").strip(),
            "area": p.area or "",
            "email": p.user.email if p.user_id else "",
            "avatar_url": _get_avatar_url(p.user) if p.user_id else "",
            "coordination": p.area or "",
            "modality": "",
            "contract_start": "",
            "contract_end": "",
            "contract_number": p.employee_number or "",
            "education_level": "",
            "undergraduate": "",
            "postgraduate": "",
        })
    for p in ContractorProfile.objects.select_related("user").order_by("full_name"):
        instructors.append({
            "role": "Contratista",
            "full_name": p.full_name,
            "document": p.document_number,
            "specialty": (p.specialty or "").strip(),
            "area": p.coordination or "",
            "email": p.user.email if p.user_id else "",
            "avatar_url": _get_avatar_url(p.user) if p.user_id else "",
            "coordination": p.coordination or "",
            "modality": p.modality or "",
            "contract_start": p.contract_start_date.strftime("%d/%m/%Y") if p.contract_start_date else "",
            "contract_end": p.contract_end_date.strftime("%d/%m/%Y") if p.contract_end_date else "",
            "contract_number": p.contract_number or "",
            "education_level": p.education_level or "",
            "undergraduate": p.undergraduate or "",
            "postgraduate": p.postgraduate or "",
        })
    instructors.sort(key=lambda x: x["full_name"])

    seen_sp: set = set()
    specialties_list = []
    for inst in instructors:
        sp = inst["specialty"]
        if sp:
            sp_norm = sp.lower()
            if sp_norm not in seen_sp:
                seen_sp.add(sp_norm)
                specialties_list.append(sp)
    specialties_list.sort(key=str.lower)

    ambiences = []
    for ambiente in Ambiente.objects.select_related("sede").order_by("sede__nombre", "nombre"):
        ambiences.append(
            {
                "id": ambiente.id,
                "sede": ambiente.sede.nombre,
                "nombre": ambiente.nombre,
                "descripcion": (ambiente.descripcion or "").strip(),
            }
        )

    sedes_list = sorted({a["sede"] for a in ambiences}, key=str.lower)

    return render(
        request,
        "ficha_config.html",
        {
            "active_tab": "crear",
            "ficha": ficha,
            "results": results,
            "trimestres": trimestres,
            "instructors": instructors,
            "specialties_list": specialties_list,
            "ambiences": ambiences,
            "sedes_list": sedes_list,
        },
    )


@login_required
@require_http_methods(["POST"])
def save_schedule_view(request, pk):
    """Guardar estado del horario en BD (caché)"""
    if not _is_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)
    try:
        ficha = Ficha.objects.get(pk=pk)
        data = json.loads(request.body)
        state_str = data.get("state", "[]")
        state = json.loads(state_str) if isinstance(state_str, str) else state_str
        ficha.schedule_state = state
        ficha.save(update_fields=["schedule_state", "updated_at"])
        return JsonResponse({"ok": True})
    except Ficha.DoesNotExist:
        return JsonResponse({"error": "Ficha no encontrada"}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido"}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def load_schedule_view(request, pk):
    """Cargar estado del horario desde BD (caché)"""
    if not _is_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)
    try:
        ficha = Ficha.objects.get(pk=pk)
        state = ficha.schedule_state or []
        return JsonResponse({"state": json.dumps(state)})
    except Ficha.DoesNotExist:
        return JsonResponse({"error": "Ficha no encontrada"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def clear_schedule_view(request, pk):
    """Limpiar caché del horario de la ficha"""
    if not _is_admin(request.user):
        return JsonResponse({"error": "Sin permisos"}, status=403)
    try:
        ficha = Ficha.objects.get(pk=pk)
        ficha.schedule_state = []
        ficha.save(update_fields=["schedule_state", "updated_at"])
        return JsonResponse({"ok": True})
    except Ficha.DoesNotExist:
        return JsonResponse({"error": "Ficha no encontrada"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@never_cache
@login_required
@require_http_methods(["GET", "POST"])
def fichas_create_view(request):
    if not _is_admin(request.user):
        logout(request)
        messages.error(request, "Tu sesion no tiene permisos admin.")
        return redirect("login")

    form_data = {
        "codigo": "",
        "matrix_id": "",
        "estado": "",
        "cdf_trimestres": "",
        "jornada": "",
        "fecha_inicio_lectiva": "",
        "fecha_fin_lectiva": "",
    }

    if request.method == "POST":
        action = request.POST.get("action", "")

        if action == "create_ficha":
            codigo = request.POST.get("codigo", "").strip()
            matrix_id = request.POST.get("matrix_id", "").strip()
            estado = request.POST.get("estado", "").strip()
            cdf_trimestres = request.POST.get("cdf_trimestres", "").strip()
            jornada = request.POST.get("jornada", "").strip()
            fecha_inicio_raw = request.POST.get("fecha_inicio_lectiva", "").strip()
            fecha_fin_raw = request.POST.get("fecha_fin_lectiva", "").strip()
            matrix = MatrixUpload.objects.filter(pk=matrix_id).first() if matrix_id.isdigit() else None

            form_data["codigo"] = codigo
            form_data["matrix_id"] = matrix_id
            form_data["estado"] = estado
            form_data["cdf_trimestres"] = cdf_trimestres
            form_data["jornada"] = jornada
            form_data["fecha_inicio_lectiva"] = fecha_inicio_raw
            form_data["fecha_fin_lectiva"] = fecha_fin_raw

            fecha_inicio = None
            fecha_fin = None
            fecha_error = False
            if fecha_inicio_raw:
                try:
                    fecha_inicio = dt_date.fromisoformat(fecha_inicio_raw)
                except ValueError:
                    fecha_error = True
            if fecha_fin_raw:
                try:
                    fecha_fin = dt_date.fromisoformat(fecha_fin_raw)
                except ValueError:
                    fecha_error = True

            matrix_max_trimester = 0
            if matrix:
                for competency in matrix.competencies.prefetch_related("results").all():
                    for result in competency.results.all():
                        try:
                            value = int(result.trimester)
                            if value > matrix_max_trimester:
                                matrix_max_trimester = value
                        except (ValueError, TypeError):
                            continue

            if not codigo:
                messages.error(request, "Debes ingresar el codigo de la ficha.")
            elif Ficha.objects.filter(codigo__iexact=codigo).exists():
                messages.error(request, "Ese codigo de ficha ya existe.")
            elif not matrix:
                messages.error(request, "Debes seleccionar un programa valido.")
            elif estado not in {Ficha.ESTADO_ABIERTA, Ficha.ESTADO_CERRADA}:
                messages.error(request, "Debes seleccionar el estado de la ficha.")
            elif cdf_trimestres and cdf_trimestres not in {"5", "6"}:
                messages.error(request, "CDF solo admite 5 o 6 trimestres cuando se seleccione.")
            elif jornada not in {item[0] for item in Ficha.JORNADA_CHOICES}:
                messages.error(request, "Debes seleccionar la jornada.")
            elif not fecha_inicio or not fecha_fin or fecha_error:
                messages.error(request, "Debes ingresar fechas lectivas validas.")
            elif fecha_fin < fecha_inicio:
                messages.error(request, "La fecha fin lectiva no puede ser menor a la fecha inicio.")
            elif cdf_trimestres and matrix_max_trimester < int(cdf_trimestres):
                messages.error(
                    request,
                    "El programa seleccionado tiene menos trimestres que el CDF elegido.",
                )
            else:
                cdf_value = int(cdf_trimestres) if cdf_trimestres else None
                Ficha.objects.create(
                    codigo=codigo,
                    matrix=matrix,
                    version_programa=matrix.version or "",
                    estado=estado,
                    cdf_trimestres=cdf_value,
                    jornada=jornada,
                    fecha_inicio_lectiva=fecha_inicio,
                    fecha_fin_lectiva=fecha_fin,
                    created_by=request.user,
                )
                messages.success(request, "Ficha registrada correctamente.")
                return redirect("fichas-panel")

    matrices = list(
        MatrixUpload.objects.prefetch_related("competencies__results").all().order_by("program_name", "project_code", "-created_at")
    )

    matrices_with_trimester = []
    for matrix in matrices:
        max_trimester = 0
        for competency in matrix.competencies.all():
            for result in competency.results.all():
                try:
                    value = int(result.trimester)
                    if value > max_trimester:
                        max_trimester = value
                except (ValueError, TypeError):
                    continue

        matrices_with_trimester.append(
            {
                "obj": matrix,
                "max_trimester": max_trimester,
            }
        )

    return render(
        request,
        "fichas_create.html",
        {
            "active_tab": "crear",
            "matrices": matrices_with_trimester,
            "form_data": form_data,
        },
    )
