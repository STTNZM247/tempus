from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any
import unicodedata

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.db import transaction

from openpyxl import load_workbook

from tempus.models import ContractorProfile


@dataclass
class UsersImportSummary:
    rows_processed: int
    skipped_rows: int
    created_users: int
    updated_users: int
    contractors: list[dict[str, Any]]


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, str):
        return " ".join(value.replace("\n", " ").split()).strip()

    if isinstance(value, float):
        rounded = round(value, 4)
        if rounded.is_integer():
            return str(int(rounded))
        return f"{rounded:.4f}".rstrip("0").rstrip(".")

    return str(value).strip()


def _match_text(value: str) -> str:
    lowered = value.lower().strip()
    normalized = unicodedata.normalize("NFKD", lowered)
    without_accents = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return " ".join(without_accents.split())


def _safe_email(value: str) -> str:
    email = value.strip().lower()
    return email if "@" in email else ""


def _parse_excel_date(value: Any) -> date | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = _normalize_text(value)
    if not text:
        return None

    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.date()
        except ValueError:
            continue

    return None


def _build_username_seed(email: str, document_number: str, full_name: str) -> str:
    if email:
        return email.split("@", 1)[0][:30]

    if document_number:
        return f"ctr_{document_number}"[:30]

    compact = "".join(ch for ch in full_name.lower() if ch.isalnum())
    return (compact or "contratista")[:30]


def _ensure_unique_username(base_username: str) -> str:
    User = get_user_model()

    username = base_username[:150] or "contratista"
    if not User.objects.filter(username=username).exists():
        return username

    suffix = 2
    while True:
        candidate = f"{username[:140]}_{suffix}"
        if not User.objects.filter(username=candidate).exists():
            return candidate
        suffix += 1


def _validate_users_headers(sheet) -> None:
    expected_header_tokens_by_col: dict[int, tuple[str, ...]] = {
        1: ("no",),
        2: ("contrato",),
        3: ("nombre", "apellidos"),
        4: ("documento", "identidad"),
        5: ("correo", "sena"),
        6: ("nivel", "formacion"),
        7: ("pregrado",),
        8: ("posgrado",),
        9: ("coordinacion",),
        10: ("modalidad", "formacion"),
        11: ("especialidad",),
        12: ("fecha", "inicio", "contrato"),
        13: ("fecha", "finalizacion", "contrato"),
    }

    if sheet.max_column < 13:
        raise ValueError("Error: archivo o datos incompatibles para cargar contratistas.")

    invalid_columns: list[str] = []
    for col_idx, tokens in expected_header_tokens_by_col.items():
        header_text = _match_text(_normalize_text(sheet.cell(row=1, column=col_idx).value))
        if not header_text or any(token not in header_text for token in tokens):
            invalid_columns.append(chr(64 + col_idx))

    if invalid_columns:
        raise ValueError("Error: archivo o datos incompatibles para cargar contratistas.")


def parse_users_workbook(file_bytes: bytes) -> UsersImportSummary:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    _validate_users_headers(sheet)

    header_row = 1
    headers_by_col: dict[int, str] = {}
    for col_idx in range(1, sheet.max_column + 1):
        raw_header = _normalize_text(sheet.cell(row=header_row, column=col_idx).value)
        if raw_header:
            headers_by_col[col_idx] = _match_text(raw_header)

    def col_for(*tokens: str) -> int | None:
        for idx, value in headers_by_col.items():
            if all(token in value for token in tokens):
                return idx
        return None

    columns = {
        "contract_number": col_for("contrato") or col_for("n", "contrato"),
        "full_name": col_for("nombre", "apellidos") or col_for("nombre"),
        "document_number": col_for("documento", "identidad") or col_for("documento"),
        "sena_email": col_for("correo") or col_for("email"),
        "education_level": col_for("nivel", "formacion"),
        "undergraduate": col_for("pregrado"),
        "postgraduate": col_for("posgrado") or col_for("postgrado"),
        "coordination": col_for("coordinacion") or col_for("coordina"),
        "modality": col_for("modalidad"),
        "specialty": col_for("especialidad"),
        "contract_start_date": col_for("fecha", "inicio"),
        "contract_end_date": col_for("fecha", "finalizacion") or col_for("fecha", "final"),
    }

    contractors: list[dict[str, Any]] = []
    skipped_rows = 0
    rows_processed = 0

    for row_idx in range(2, sheet.max_row + 1):
        row_data: dict[str, Any] = {"source_row": row_idx}
        for field_name, col_idx in columns.items():
            if not col_idx:
                row_data[field_name] = ""
                continue

            raw_value = sheet.cell(row=row_idx, column=col_idx).value
            if field_name in {"contract_start_date", "contract_end_date"}:
                row_data[field_name] = _parse_excel_date(raw_value)
            else:
                row_data[field_name] = _normalize_text(raw_value)

        if not row_data["contract_number"] and not row_data["document_number"] and not row_data["full_name"]:
            skipped_rows += 1
            continue

        rows_processed += 1
        row_data["sena_email"] = _safe_email(row_data["sena_email"])
        contractors.append(row_data)

    return UsersImportSummary(
        rows_processed=rows_processed,
        skipped_rows=skipped_rows,
        created_users=0,
        updated_users=0,
        contractors=contractors,
    )


def save_users_summary(summary: dict[str, Any]) -> tuple[int, int]:
    contractors = summary.get("contractors", [])
    if not contractors:
        raise ValueError("No hay contratistas para guardar.")

    User = get_user_model()
    group, _ = Group.objects.get_or_create(name="Contratista")

    created_users = 0
    updated_users = 0

    with transaction.atomic():
        for contractor in contractors:
            contract_number = contractor.get("contract_number", "").strip()
            document_number = contractor.get("document_number", "").strip()
            full_name = contractor.get("full_name", "").strip() or "Sin nombre"
            sena_email = contractor.get("sena_email", "").strip().lower()

            if not contract_number or not document_number:
                continue

            profile = (
                ContractorProfile.objects.select_related("user")
                .filter(document_number=document_number)
                .first()
            )
            if profile is None:
                profile = (
                    ContractorProfile.objects.select_related("user")
                    .filter(contract_number=contract_number)
                    .first()
                )

            if profile:
                user = profile.user
                updated_users += 1
            else:
                user = None
                if sena_email:
                    user = User.objects.filter(email__iexact=sena_email).first()
                if user is None:
                    seed = _build_username_seed(sena_email, document_number, full_name)
                    username = _ensure_unique_username(seed)
                    user = User.objects.create(username=username, email=sena_email)
                    user.set_unusable_password()
                    user.is_staff = False
                    user.is_superuser = False
                    user.first_name = full_name[:150]
                    user.save()
                    created_users += 1
                else:
                    updated_users += 1

            if user.email != sena_email:
                user.email = sena_email

            if user.first_name != full_name[:150]:
                user.first_name = full_name[:150]

            if user.is_staff:
                user.is_staff = False

            if user.is_superuser:
                user.is_superuser = False

            user.save()
            user.groups.add(group)

            if profile is None:
                profile = ContractorProfile(document_number=document_number)

            profile.user = user
            profile.contract_number = contract_number
            profile.document_number = document_number
            profile.full_name = full_name
            profile.sena_email = sena_email
            profile.education_level = contractor.get("education_level", "")
            profile.undergraduate = contractor.get("undergraduate", "")
            profile.postgraduate = contractor.get("postgraduate", "")
            profile.coordination = contractor.get("coordination", "")
            profile.modality = contractor.get("modality", "")
            profile.specialty = contractor.get("specialty", "")
            profile.contract_start_date = _parse_excel_date(contractor.get("contract_start_date"))
            profile.contract_end_date = _parse_excel_date(contractor.get("contract_end_date"))
            profile.source_row = contractor.get("source_row", 0) or 0
            profile.save()

    return created_users, updated_users
