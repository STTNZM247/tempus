from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any
import unicodedata
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from django.db import transaction

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from tempus.models import MatrixCompetency, MatrixResult, MatrixUpload


@dataclass
class MatrixImportSummary:
    project_code: str | None
    project_description: str | None
    program_code: str | None
    program_name: str | None
    level: str | None
    version: str | None
    competencies: list[dict[str, Any]]
    rows_processed: int
    skipped_rows: int


def save_matrix_summary(summary: dict[str, Any], uploaded_by: Any) -> MatrixUpload:
    project_code = summary.get("project_code")
    if not project_code:
        raise ValueError("La matriz no trae codigo de proyecto.")

    if MatrixUpload.objects.filter(project_code=project_code).exists():
        raise ValueError("Ya existe una matriz registrada con ese codigo de proyecto.")

    competencies = summary.get("competencies", [])
    result_count = sum(len(competency.get("results", [])) for competency in competencies)

    with transaction.atomic():
        matrix = MatrixUpload.objects.create(
            project_code=project_code,
            project_description=summary.get("project_description") or "",
            program_code=summary.get("program_code") or "",
            program_name=summary.get("program_name") or "",
            level=summary.get("level") or "",
            version=summary.get("version") or "",
            competency_count=len(competencies),
            result_count=result_count,
            uploaded_by=uploaded_by,
        )

        for competency_index, competency_data in enumerate(competencies, start=1):
            competency = MatrixCompetency.objects.create(
                matrix=matrix,
                code=competency_data.get("code") or "",
                name=competency_data.get("name") or "",
                hours=competency_data.get("hours") or "",
                display_order=competency_index,
            )

            for result_index, result_data in enumerate(competency_data.get("results", []), start=1):
                MatrixResult.objects.create(
                    competency=competency,
                    name=result_data.get("name") or "",
                    hours_max=result_data.get("hours_max") or "",
                    hours_min=result_data.get("hours_min") or "",
                    trimester=result_data.get("trimester") or "",
                    weekly_hours=result_data.get("weekly_hours") or "",
                    trimester_hours=result_data.get("trimester_hours") or "",
                    display_order=result_index,
                )

    return matrix


def _normalize_value(value: Any) -> str | None:
    if value is None:
        return None

    if isinstance(value, str):
        cleaned = " ".join(value.replace("\n", " ").split())
        return cleaned or None

    if isinstance(value, float):
        rounded = round(value, 4)
        if rounded.is_integer():
            return str(int(rounded))
        return f"{rounded:.4f}".rstrip("0").rstrip(".")

    return str(value).strip() or None


def _to_match_text(value: str | None) -> str:
    if not value:
        return ""

    lowered = value.lower()
    normalized = unicodedata.normalize("NFKD", lowered)
    without_accents = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return " ".join(without_accents.split())


def _round_int_string(value: str | None) -> str | None:
    if value is None:
        return None

    normalized = value.replace(",", ".").strip()
    if not normalized:
        return None

    try:
        number = Decimal(normalized)
    except InvalidOperation:
        return value

    return str(int(number.quantize(Decimal("1"), rounding=ROUND_HALF_UP)))


def _autofill_hours_min(hours_min: str | None, hours_max: str | None) -> str | None:
    rounded_min = _round_int_string(hours_min)
    if rounded_min is not None:
        return rounded_min

    if hours_max is None:
        return None

    normalized_max = hours_max.replace(",", ".").strip()
    try:
        max_number = Decimal(normalized_max)
    except InvalidOperation:
        return None

    computed_min = max_number * Decimal("0.8")
    return str(int(computed_min.quantize(Decimal("1"), rounding=ROUND_HALF_UP)))


def _merged_value(sheet: Worksheet, row: int, column: int) -> Any:
    cell = sheet.cell(row=row, column=column)
    if cell.value is not None:
        return cell.value

    for merged_range in sheet.merged_cells.ranges:
        if row in range(merged_range.min_row, merged_range.max_row + 1) and column in range(merged_range.min_col, merged_range.max_col + 1):
            return sheet.cell(merged_range.min_row, merged_range.min_col).value

    return None


def parse_matrix_workbook(file_bytes: bytes) -> MatrixImportSummary:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sheet = workbook.active

    active_program: dict[str, Any] = {
        "project_code": None,
        "project_description": None,
        "program_code": None,
        "program_name": None,
        "level": None,
        "version": None,
    }
    active_competency: dict[str, Any] | None = None
    competencies: list[dict[str, Any]] = []
    skipped_rows = 0
    rows_processed = 0
    header_markers = {
        "codigo del proyecto",
        "descripcion del proyecto",
        "codigo del programa",
        "programa de formacion",
        "nombre de la competencia laboral",
        "nombre de la competencia",
        "resultado de aprendizaje",
        "duracion de la competencia en horas",
        "duracion horas por resultado",
        "duracion horas por resultado de aprendizaje",
        "trimestre a programar",
    }

    start_data_row = 2
    for row_index in range(start_data_row, sheet.max_row + 1):
        program_data = {
            "project_code": _normalize_value(_merged_value(sheet, row_index, 1)),
            "project_description": _normalize_value(_merged_value(sheet, row_index, 2)),
            "program_code": _normalize_value(_merged_value(sheet, row_index, 3)),
            "program_name": _normalize_value(_merged_value(sheet, row_index, 4)),
            "level": _normalize_value(_merged_value(sheet, row_index, 5)),
            "version": _normalize_value(_merged_value(sheet, row_index, 6)),
        }

        competency_name = _normalize_value(_merged_value(sheet, row_index, 7))
        competency_code = _normalize_value(_merged_value(sheet, row_index, 8))
        competency_hours = _normalize_value(_merged_value(sheet, row_index, 9))
        result_name = _normalize_value(_merged_value(sheet, row_index, 10))
        result_hours_max = _normalize_value(_merged_value(sheet, row_index, 11))
        result_hours_min = _normalize_value(_merged_value(sheet, row_index, 12))
        result_trimester = _normalize_value(_merged_value(sheet, row_index, 13))
        result_weekly_hours = _normalize_value(_merged_value(sheet, row_index, 14))
        result_trimester_hours = _normalize_value(_merged_value(sheet, row_index, 15))

        row_text = " ".join(
            _to_match_text(value)
            for value in [
                program_data["project_code"],
                program_data["project_description"],
                program_data["program_code"],
                program_data["program_name"],
                program_data["level"],
                program_data["version"],
                competency_name,
                competency_code,
                competency_hours,
                result_name,
                result_hours_max,
                result_hours_min,
                result_trimester,
                result_weekly_hours,
                result_trimester_hours,
            ]
            if value
        )

        if any(marker in row_text for marker in header_markers):
            skipped_rows += 1
            continue

        for key, value in program_data.items():
            if value:
                active_program[key] = value

        if competency_name and (active_competency is None or active_competency["name"] != competency_name or active_competency["code"] != competency_code):
            active_competency = {
                "code": competency_code,
                "name": competency_name,
                "hours": competency_hours,
                "results": [],
            }
            competencies.append(active_competency)

        has_meaningful_row = any(
            [
                active_program["program_code"],
                competency_name,
                result_name,
            ]
        )

        if not has_meaningful_row:
            skipped_rows += 1
            continue

        if result_name and active_competency is not None:
            active_competency["results"].append(
                {
                    "name": result_name,
                    "hours_max": result_hours_max,
                    "hours_min": _autofill_hours_min(result_hours_min, result_hours_max),
                    "trimester": result_trimester,
                    "weekly_hours": result_weekly_hours,
                    "trimester_hours": result_trimester_hours,
                }
            )
            rows_processed += 1
        elif competency_name:
            rows_processed += 1
        else:
            skipped_rows += 1

    return MatrixImportSummary(
        project_code=active_program["project_code"],
        project_description=active_program["project_description"],
        program_code=active_program["program_code"],
        program_name=active_program["program_name"],
        level=active_program["level"],
        version=active_program["version"],
        competencies=competencies,
        rows_processed=rows_processed,
        skipped_rows=skipped_rows,
    )
