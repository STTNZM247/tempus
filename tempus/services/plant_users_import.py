from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.db import transaction

from openpyxl import load_workbook

from tempus.models import PlantProfile
from tempus.services.users_import import _ensure_unique_username, _match_text, _normalize_text


@dataclass
class PlantUsersImportSummary:
    rows_processed: int
    skipped_rows: int
    created_users: int
    updated_users: int
    instructors: list[dict[str, Any]]


def _build_username_seed(document_number: str, full_name: str) -> str:
    if document_number:
        return f"pl_{document_number}"[:30]

    compact = "".join(ch for ch in full_name.lower() if ch.isalnum())
    return (compact or "planta")[:30]


def _find_header_row(sheet) -> int:
    target_tokens = ("nombre", "cedula", "area", "estudios")

    for row_idx in range(1, min(sheet.max_row, 8) + 1):
        matches = 0
        for col_idx in range(1, sheet.max_column + 1):
            cell_value = _match_text(_normalize_text(sheet.cell(row=row_idx, column=col_idx).value))
            if any(token in cell_value for token in target_tokens):
                matches += 1

        if matches >= 3:
            return row_idx

    return 1


def _validate_plant_headers(sheet) -> int:
    header_row = _find_header_row(sheet)

    expected_header_tokens_by_col: dict[int, tuple[str, ...]] = {
        1: ("n",),
        2: ("nombre", "apellidos"),
        3: ("cedula",),
        4: ("area",),
        5: ("estudios",),
    }

    if sheet.max_column < 5:
        raise ValueError("Error: archivo o datos incompatibles para cargar planta.")

    for col_idx, tokens in expected_header_tokens_by_col.items():
        header_text = _match_text(_normalize_text(sheet.cell(row=header_row, column=col_idx).value))
        if not header_text or any(token not in header_text for token in tokens):
            raise ValueError("Error: archivo o datos incompatibles para cargar planta.")

    return header_row


def parse_plant_users_workbook(file_bytes: bytes) -> PlantUsersImportSummary:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sheet = workbook.active

    header_row = _validate_plant_headers(sheet)
    headers_by_col: dict[int, str] = {}
    for col_idx in range(1, sheet.max_column + 1):
        raw_header = _normalize_text(sheet.cell(row=header_row, column=col_idx).value)
        if raw_header:
            headers_by_col[col_idx] = _match_text(raw_header)

    def col_for(*tokens: str) -> int | None:
        for idx, value in headers_by_col.items():
            if all(token in value for token in tokens):
                return idx
        return None

    columns = {
        "employee_number": col_for("n") or col_for("numero"),
        "full_name": col_for("nombre", "apellidos") or col_for("nombre"),
        "document_number": col_for("cedula") or col_for("documento"),
        "area": col_for("area"),
        "studies": col_for("estudios"),
    }

    instructors: list[dict[str, Any]] = []
    skipped_rows = 0
    rows_processed = 0

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        row_data: dict[str, Any] = {"source_row": row_idx}
        for field_name, col_idx in columns.items():
            if not col_idx:
                row_data[field_name] = ""
                continue
            row_data[field_name] = _normalize_text(sheet.cell(row=row_idx, column=col_idx).value)

        if not row_data["full_name"] and not row_data["document_number"]:
            skipped_rows += 1
            continue

        rows_processed += 1
        instructors.append(row_data)

    return PlantUsersImportSummary(
        rows_processed=rows_processed,
        skipped_rows=skipped_rows,
        created_users=0,
        updated_users=0,
        instructors=instructors,
    )


def save_plant_users_summary(summary: dict[str, Any]) -> tuple[int, int]:
    instructors = summary.get("instructors", [])
    if not instructors:
        raise ValueError("No hay instructores de planta para guardar.")

    User = get_user_model()
    group, _ = Group.objects.get_or_create(name="Planta")

    created_users = 0
    updated_users = 0

    with transaction.atomic():
        for instructor in instructors:
            document_number = instructor.get("document_number", "").strip()
            full_name = instructor.get("full_name", "").strip() or "Sin nombre"
            employee_number = instructor.get("employee_number", "").strip()

            if not document_number:
                continue

            profile = (
                PlantProfile.objects.select_related("user")
                .filter(document_number=document_number)
                .first()
            )

            if profile:
                user = profile.user
                updated_users += 1
            else:
                user = None
                seed = _build_username_seed(document_number, full_name)
                username = _ensure_unique_username(seed)
                user = User.objects.create(username=username, email="")
                user.set_unusable_password()
                user.is_staff = False
                user.is_superuser = False
                user.first_name = full_name[:150]
                user.save()
                created_users += 1

            if user.first_name != full_name[:150]:
                user.first_name = full_name[:150]
                user.save(update_fields=["first_name"])

            user.groups.add(group)

            if profile is None:
                profile = PlantProfile(document_number=document_number)

            profile.user = user
            profile.employee_number = employee_number
            profile.document_number = document_number
            profile.full_name = full_name
            profile.area = instructor.get("area", "")
            profile.studies = instructor.get("studies", "")
            profile.source_row = instructor.get("source_row", 0) or 0
            profile.save()

    return created_users, updated_users
